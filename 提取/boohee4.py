from urllib import parse,request
from bs4 import BeautifulSoup
import pypinyin
import openpyxl

def yingyangsu():
    t1 = openpyxl.load_workbook(filename="C:/Users/meridian/Desktop/提取练习/主辅料信息.xlsx")
    e1 = t1.active
    l3 = []
    for i in e1['A']:
        try:
            s = ''
            pinlist = pypinyin.pinyin(i.value, style=pypinyin.NORMAL)
            for j in pinlist:
                s += ''.join(j)

            url = 'http://www.boohee.com/shiwu/'

            url = url + s

            rsp = request.urlopen(url=url)

            html = rsp.read()

            html = html.decode()

            soup = BeautifulSoup(html,'lxml')

            divs = soup.find_all(class_='nutr-tag margin10')

            # print(divs[0].get_text())
            str_ = divs[0].get_text()
            str_ = str_.replace('\n', '^').replace('\n', '^')
            l1 = str_.split('^')
            l2 =[]
            l2.append(i.value)
            for j in l1:
                if j !='' and j[0] != "营":

                    l2.append(j)

            l3.append(l2)
            print(l2)

        except Exception as e:
            print(i.value,'找不到')
    tq = openpyxl.Workbook()
    Sheet1 = tq.active
    Sheet1.title = "已提取"
    # l5=list()
    print("开始写入")

    # 行数很少时可以使用
    # Sheet1.append(l4)

    for i in l3:
        Sheet1.append(i)

    tq.save('C:/Users/meridian/Desktop/提取练习/已处理.xlsx')
def duliangdanwei():

    t1 = openpyxl.load_workbook(filename="C:/Users/meridian/Desktop/提取练习/主辅料信息.xlsx")
    e1 = t1.active

    l4 = []
    for i in e1['A']:
        try:
            s = ''
            pinlist = pypinyin.pinyin(i.value, style=pypinyin.NORMAL)
            for j in pinlist:
                s += ''.join(j)

            url = 'http://www.boohee.com/shiwu/'

            url = url + s

            rsp = request.urlopen(url=url)

            html = rsp.read()

            html = html.decode()

            soup = BeautifulSoup(html, 'lxml')

            divs = soup.find_all(class_='widget-unit')
            str_ = divs[0].get_text()
            str_ = str_.replace('\n', '^').replace('\n', '^').replace(' ','^')
            l1 = str_.split('^')
            l2 =[]
            l3 = []
            for j in l1:
                if j != '':
                    l2.append(j)

            l3.append(i.value)

            l3.append(l2[3])
            l3.append(l2[4])

            l4.append(l3)


            print(l3)

        except Exception as e:
            print(i.value, '找不到')
    tq = openpyxl.Workbook()
    Sheet1 = tq.active
    Sheet1.title = "已提取"
    # l5=list()
    print("开始写入")

    # 行数很少时可以使用
    # Sheet1.append(l4)

    for i in l4:
        Sheet1.append(i)

    tq.save('C:/Users/meridian/Desktop/提取练习/已处理.xlsx')


def yuanliao():
    t1 = openpyxl.load_workbook(filename="C:/Users/meridian/Desktop/提取练习/主辅料信息.xlsx")
    e1 = t1.active
    l3 = []
    for i in e1['A']:
        try:
            s = ''
            pinlist = pypinyin.pinyin(i.value, style=pypinyin.NORMAL)
            for j in pinlist:
                s += ''.join(j)

            url = 'http://www.boohee.com/shiwu/'

            url = url + s

            rsp = request.urlopen(url=url)

            html = rsp.read()

            html = html.decode()

            soup = BeautifulSoup(html, 'lxml')

            divs = soup.find_all(class_='widget-more')

            str_ = divs[0].get_text()
            str_ = str_.replace('\n', '^').replace('\n', '^').replace(' ', '^')
            l1 = str_.split('^')
            l2 = []
            l2.append(i.value)
            for j in l1:
                if j != '' :
                    l2.append(j)
                if '做法' in j or '详细' in j :
                    break


            print(l2)


            l3.append(l2)
        except Exception as e:
            print(i.value,"找不到")
        tq = openpyxl.Workbook()
        Sheet1 = tq.active
        Sheet1.title = "已提取"



        for i in l3:
            Sheet1.append(i)

        tq.save('C:/Users/meridian/Desktop/提取练习/已处理.xlsx')



if __name__ == '__main__':

        # yingyangsu()


    # duliangdanwei()
    yuanliao()

