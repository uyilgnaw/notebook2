import openpyxl
from multiprocessing import Process

def Demo():
    e1 = openpyxl.load_workbook("C:/Users/meridian/Desktop/提取练习/待处理.xlsx")
    # e1 = openpyxl.load_workbook("C:/Users/meridian/Desktop/提取练习/待处理.xlsx")
    w1 = e1.active
    l1 = []



    # l1为所有行的汇总列表

    for i in w1['A']:
        lll = []
        i = i.value

        j = i.split('；')
        # print(j)
        #
        j.append(i)

        for m in j:

            if m != '':
                lll.append(m)

        l1.append(lll)
    # print(l1)
    l2 = []



    for i in l1:
        l4 = []
        l3 = []

        l4.append(i[-1])
        i.pop()
        for j in i:
           if '脾' in j:
                l3.append(j)

        l4.append('。'.join(l3))
        #     l4.append('。'.join(l5))
        l2.append(l4)

    e2 = openpyxl.Workbook()
    w2 = e2.active
    w2.title = "已提取"
    m = 0
    for i in l2:
        w2.append(i)
        m = m + 1
        print("已经提取了{0}行".format(m))

    print("提取完成!")
    print("保存中...")
    e2.save("C:/Users/meridian/Desktop/提取练习/已提取(分号分隔).xlsx")
    # e2.save("C:/Users/meridian/Desktop/提取练习/乳腺超声已提取.xlsx")

if __name__ == '__main__':
    p = Process(target=Demo)
    p.daemon = True
    p.start()
    p.join()
    print("保存完成")
